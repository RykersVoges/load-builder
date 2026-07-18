"""
Load-Building Prototype
========================
Implements the logic described in the Instruction document against the
real data in "Claude Input File.xlsx".

Assembly v2 adds an "adaptive" grouping mode: instead of pre-slicing
customers into fixed-width direction buckets (which can strand volume that
sits just across a bucket boundary), it seeds each new load with the
oldest-due pending order, then greedily pulls in the geographically nearest
other pending orders (by bearing, not a fixed bucket) until the truck is as
full as it can be. It also does a best-effort bin-covering scan (skips an
order that doesn't fit and keeps checking smaller ones after it, instead of
stopping at the first that doesn't fit).

Packing v2 uses a skyline (shelf) algorithm per trailer width-slot instead
of a single-item-per-bay model, so leftover shelf space beside a shorter
stacked bundle stays open for a later, shorter bundle to fill.
"""
import math
import os
from collections import defaultdict

import openpyxl

LB_VERSION = "v40"
INPUT_FILE = "Claude Input File.xlsx"

DIRECTION_DEGREES = 30
GROUP_MODE = "adaptive"   # "adaptive" (recommended), "direction", or "province"
MAX_DROPS_PER_LOAD = 10

# Hard cap on how far (in degrees of bearing from the site) adaptive mode
# will reach to fill a truck. Bearing alone doesn't capture distance, so
# without this cap a truck can end up combining freight in genuinely
# unrelated directions (e.g. Limpopo and Cape Town can share a similar-
# looking bearing from a KZN site purely by geographic coincidence) just
# because nothing closer was left to fill the remaining space. With the
# cap, a load ships under-full rather than reaching that far.
MAX_CORRIDOR_SPREAD_DEG = 60

# A truck is only dispatched if the batch reaches at least ONE of these
# utilisation thresholds (percent of that truck's payload / cube capacity).
# The final leftover batch of freight is exempt, so nothing gets stranded.
MIN_WT_UTIL_PCT = 75.0
MIN_VOL_UTIL_PCT = 75.0

TRUCK_TYPES = {
    "34T": {
        "display_name": "34 Ton Tautliner",
        "trailers": [
            {"name": "front", "length_m": 6, "weight_cap_t": 12, "width_m": 2.5, "height_m": 2.6},
            {"name": "rear",  "length_m": 12, "weight_cap_t": 24, "width_m": 2.5, "height_m": 2.6},
        ],
        "payload_cap_t": 36, "cube_cap_m3": 117,
        "min_weight_t": 28, "min_vol_m3": 55,
        "fleet_count": 20,
    },
    "FD": {
        # Flat Deck -- placeholder single-trailer spec until the uploaded
        # Truck Dimensions tab's own "34 Ton Flat Deck" (or similar) column
        # overrides these, including display_name itself (see
        # _apply_truck_dimensions). Defaults to 0 in the fleet so it has no
        # effect unless the user actually owns some.
        "display_name": "34 Ton Flat Deck",
        "trailers": [{"name": "single", "length_m": 12.5, "weight_cap_t": 34, "width_m": 2.5, "height_m": 2.6}],
        "payload_cap_t": 34, "cube_cap_m3": 100,
        "min_weight_t": 26, "min_vol_m3": 45,
        "fleet_count": 0,
    },
    "30T": {
        "display_name": "30 Ton Tri Axle Tautliner",
        "trailers": [{"name": "single", "length_m": 18, "weight_cap_t": 30, "width_m": 2.5, "height_m": 2.6}],
        "payload_cap_t": 30, "cube_cap_m3": 117,
        "min_weight_t": 23, "min_vol_m3": 48,
        "fleet_count": 0,
    },
    "14T": {
        "display_name": "14 Ton Tautliner",
        "trailers": [{"name": "single", "length_m": 14, "weight_cap_t": 14, "width_m": 2.5, "height_m": 2.6}],
        "payload_cap_t": 14, "cube_cap_m3": 91,
        "min_weight_t": 11, "min_vol_m3": 20,
        "fleet_count": 0,
    },
    "8T": {
        "display_name": "8 Ton Tautliner",
        "trailers": [{"name": "single", "length_m": 8, "weight_cap_t": 8, "width_m": 2.5, "height_m": 2.5}],
        "payload_cap_t": 8, "cube_cap_m3": 50,
        "min_weight_t": 6, "min_vol_m3": 12,
        "fleet_count": 0,
    },
}
TRUCK_TRY_ORDER = ["34T", "FD", "30T", "14T", "8T"]


def truck_display_name(key):
    """Human-readable name for a truck type key, e.g. '34T' -> '34 Ton
    Tautliner' -- picked up from the uploaded Truck Dimensions tab's own
    column header if present, otherwise a sensible default."""
    spec = TRUCK_TYPES.get(key)
    if not spec:
        return key
    return spec.get("display_name") or key

SITE_ALIAS = {"SFP_LSM": "LSM", "SFP_WSM": "WSM"}
SITE_SWAP = {"LSM": "WSM", "WSM": "LSM"}

COMPASS_POINTS = ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE",
                  "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]


def group_label(group_value):
    """Human-readable label for a group key: a compass bearing range for
    direction mode (int bucket), the province code as-is for province mode,
    or a centroid +/- spread description for adaptive corridors."""
    if isinstance(group_value, tuple) and group_value and group_value[0] == "adaptive":
        _, centroid, spread = group_value
        compass = COMPASS_POINTS[int((centroid % 360) / 22.5) % 16]
        return "~%d° +/-%d° (%s), adaptive corridor" % (round(centroid), round(spread), compass)
    if isinstance(group_value, int):
        lo = group_value * DIRECTION_DEGREES
        hi = lo + DIRECTION_DEGREES
        compass = COMPASS_POINTS[int(((lo + hi) / 2 % 360) / 22.5) % 16]
        return "%d°-%d° (%s)" % (lo, hi, compass)
    return str(group_value)


# Ships with the app; holds the Customer Locations / SKU Bundle Dimensions /
# Site Locations / Truck Dimensions reference tabs so the daily upload only
# needs the orders tab. Resolved relative to THIS file so it's found no
# matter what folder the server happens to run from.
REFERENCE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reference_data.xlsx")
_REF_WB = None

# (sales_order_digits, sku) -> {"line_id", "bucket_ext", "line_no"} parsed
# from the optional Demand Buckets tab; used by the Transport Orders output.
DEMAND_BUCKETS = {}


def _find_sheet_or_none(wb, *keywords):
    for name in wb.sheetnames:
        n = " ".join(str(name).strip().lower().split())
        if all(k in n for k in keywords):
            return wb[name]
    return None


def _find_sheet(wb, *keywords, use_reference=True):
    """Find a tab whose name contains all the given keywords, ignoring
    case, extra spaces, and small renames. If the tab isn't in the uploaded
    workbook and it's a reference tab (customers/SKUs/sites), fall back to
    the reference_data.xlsx bundled with the app -- so the daily upload
    only ever needs to contain the orders tab."""
    ws = _find_sheet_or_none(wb, *keywords)
    if ws is not None:
        return ws
    if use_reference:
        import os
        global _REF_WB
        if _REF_WB is None and os.path.exists(REFERENCE_FILE):
            _REF_WB = openpyxl.load_workbook(REFERENCE_FILE, data_only=True)
        if _REF_WB is not None:
            ws = _find_sheet_or_none(_REF_WB, *keywords)
            if ws is not None:
                return ws
    raise ValueError(
        "Could not find a tab matching %r in the uploaded workbook%s. "
        "Tabs found in the upload: %s." % (
            " + ".join(keywords),
            " or in the app's reference data" if use_reference else "",
            ", ".join(wb.sheetnames)))


def _norm(v):
    return " ".join(str(v).strip().lower().replace("³", "3").split()) if v is not None else ""


def _f(v):
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def _site_code(v):
    """SFP-LSM / SFP_LSM / LSM all mean the Langeni mill -- normalise."""
    return str(v).strip().upper().replace("SFP-", "").replace("SFP_", "")


def _tab_columns(ws, must_have):
    """Locate the header row (first row containing every name in must_have,
    ignoring case/spacing) and return (header_row_number, {name: col_index})."""
    limit = min(12, ws.max_row)
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=limit, values_only=True), 1):
        names = [_norm(v) for v in r]
        if all(m in names for m in must_have):
            colmap = {}
            for j, n in enumerate(names):
                if n and n not in colmap:
                    colmap[n] = j
            return i, colmap
    raise ValueError('Could not find a header row containing [%s] on tab "%s".'
                     % (", ".join(must_have), ws.title))


def _apply_truck_dimensions(wb):
    """Read the Truck Dimensions tab (from the upload, else the reference
    file) and override the built-in truck specs -- lengths, weights, widths,
    heights, payload/cube caps, and bundles-across all come from the user's
    workbook, not from hard-coded defaults."""
    import re
    ws = _find_sheet_or_none(wb, "truck")
    if ws is None:
        global _REF_WB
        if _REF_WB is None and os.path.exists(REFERENCE_FILE):
            _REF_WB = openpyxl.load_workbook(REFERENCE_FILE, data_only=True)
        if _REF_WB is not None:
            ws = _find_sheet_or_none(_REF_WB, "truck")
    if ws is None:
        return
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = desc_col = None
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            if _norm(v) == "descriptor":
                hdr_i, desc_col = i, j
                break
        if hdr_i is not None:
            break
    if hdr_i is None:
        return
    col_truck = {}
    col_header_text = {}
    for j, v in enumerate(rows[hdr_i]):
        if j == desc_col or v is None:
            continue
        header_norm = _norm(v)
        if "flat" in header_norm and "FD" in TRUCK_TYPES:
            # e.g. "Flat Deck", "Flatbed" -- no leading tonnage number, so
            # match on keyword instead of the digit-prefix rule below.
            col_truck[j] = "FD"
            col_header_text[j] = str(v).strip()
            continue
        m = re.match(r"\s*(\d+)", str(v))
        if m and (m.group(1) + "T") in TRUCK_TYPES:
            col_truck[j] = m.group(1) + "T"
            col_header_text[j] = str(v).strip()
    # Use the workbook's own column header text as the display name
    # everywhere in the app -- e.g. "34 Ton Flat Deck" rather than the
    # internal key "FD" -- so the Loads table, Excel output, and PDF match
    # exactly what the user's own truck master calls each type.
    for j, key in col_truck.items():
        if col_header_text.get(j):
            TRUCK_TYPES[key]["display_name"] = col_header_text[j]
    vals = {key: {} for key in col_truck.values()}
    for r in rows[hdr_i + 1:]:
        d = _norm(r[desc_col]) if desc_col < len(r) else ""
        if not d:
            continue
        for j, key in col_truck.items():
            vals[key][d] = r[j] if j < len(r) else None

    def num(d, *words):
        for k, v in d.items():
            if all(w in k for w in words):
                f = _f(v)
                if f is not None:
                    return f
        return None

    for key, d in vals.items():
        spec = TRUCK_TYPES[key]
        L1, L2 = num(d, "length", "closest"), num(d, "length", "furthe")
        W1, W2 = num(d, "weight", "closest"), num(d, "weight", "furthe")
        payload, cubes = num(d, "payload", "tons"), num(d, "payload", "cube")
        wd1 = num(d, "width", "closest") or 2.5
        wd2 = num(d, "width", "furthe") or wd1
        h1 = num(d, "height", "closest") or 2.6
        h2 = num(d, "height", "furthe") or h1
        slots = int(num(d, "bundles placed") or 2)
        if L1 is None:
            continue
        if L2 is not None:
            spec["trailers"] = [
                {"name": "front", "length_m": L1, "weight_cap_t": W1 or payload,
                 "width_m": wd1, "height_m": h1, "width_slots": slots},
                {"name": "rear", "length_m": L2, "weight_cap_t": W2 or payload,
                 "width_m": wd2, "height_m": h2, "width_slots": slots},
            ]
        else:
            spec["trailers"] = [
                {"name": "single", "length_m": L1, "weight_cap_t": W1 or payload,
                 "width_m": wd1, "height_m": h1, "width_slots": slots}]
        if payload:
            spec["payload_cap_t"] = payload
        if cubes:
            spec["cube_cap_m3"] = cubes


def load_workbook_data(source=None):
    wb = openpyxl.load_workbook(source or INPUT_FILE, data_only=True)
    _apply_truck_dimensions(wb)

    # ---- Customer Locations: carries BOTH the customers and (in its first
    # few rows, marked Location Type = FACILITY) the loading sites/mills. ----
    ws = _find_sheet(wb, "customer", "location")
    hdr, cm = _tab_columns(ws, ["location code", "lat", "lon"])
    customers, sites = {}, {}
    for r in ws.iter_rows(min_row=hdr + 1, max_row=ws.max_row, values_only=True):
        def g(*names):
            for n in names:
                j = cm.get(n)
                if j is not None and j < len(r):
                    return r[j]
            return None
        code = g("location code")
        lat, lon = _f(g("lat")), _f(g("lon"))
        if code is None or lat is None or lon is None:
            continue
        code_s = str(code).strip()
        desc = g("customer description", "description")
        if _norm(g("location type")) == "facility" or code_s.upper().startswith("SFP"):
            sites[_site_code(code_s)] = {"name": desc, "lat": lat, "lon": lon}
        else:
            customers[code_s] = {
                "area": g("area"), "city": g("city"), "desc": desc,
                "lat": lat, "lon": lon,
                "province": g("province name", "province", "prov"),
            }

    # ---- Site Locations tab is now OPTIONAL: if present (in the upload or
    # the reference file), it fills in any site the FACILITY rows didn't. ----
    ws = _find_sheet_or_none(wb, "site", "location")
    if ws is None:
        import os
        global _REF_WB
        if _REF_WB is None and os.path.exists(REFERENCE_FILE):
            _REF_WB = openpyxl.load_workbook(REFERENCE_FILE, data_only=True)
        if _REF_WB is not None:
            ws = _find_sheet_or_none(_REF_WB, "site", "location")
    if ws is not None:
        try:
            hdr, cm = _tab_columns(ws, ["site code", "lat", "lon"])
            for r in ws.iter_rows(min_row=hdr + 1, max_row=ws.max_row, values_only=True):
                code = r[cm["site code"]] if cm["site code"] < len(r) else None
                lat, lon = _f(r[cm["lat"]]), _f(r[cm["lon"]])
                if code is None or lat is None or lon is None:
                    continue
                name_j = cm.get("site")
                sites.setdefault(_site_code(code), {
                    "name": r[name_j] if name_j is not None else "", "lat": lat, "lon": lon})
        except ValueError:
            pass
    if not sites:
        raise ValueError(
            "No loading sites found. Expected FACILITY rows at the top of the "
            "Customer Locations tab, or a Site Locations tab.")

    # ---- SKU Bundle Dimensions: header-name based; prefers the
    # "incl dunnage" columns when present. ----
    ws = _find_sheet(wb, "sku")
    hdr, cm = _tab_columns(ws, ["sku code"])

    def _col_like(*words):
        for key, j in cm.items():
            if all(w in key for w in words):
                return j
        return None

    c_code = cm["sku code"]
    c_len = _col_like("unit length")
    c_h_inc, c_h = _col_like("bundle height", "inc"), _col_like("bundle height")
    c_w = _col_like("bundle width")
    c_m3_inc, c_m3 = _col_like("bundle cube", "inc"), _col_like("bundle cube")
    c_kg_inc, c_kg = _col_like("bundle kg", "inc"), _col_like("bundle kg")
    # True payload weight = Unit Weight x Qty (NOP) + dunnage. The master's
    # own "Bundle KG" column is density x OUTER bundle volume (including the
    # air between boards), which overstates the timber's real weight ~15-20%.
    c_uw = _col_like("unit weight")
    c_qty = _col_like("qty (nop)") or _col_like("qty nop")
    c_dun = cm.get("dunnage kg")

    def _val(r, j_pref, j_base=None):
        v = r[j_pref] if j_pref is not None and j_pref < len(r) else None
        if v in (None, "", 0) and j_base is not None and j_base < len(r):
            v = r[j_base]
        return v

    skus = {}
    for r in ws.iter_rows(min_row=hdr + 1, max_row=ws.max_row, values_only=True):
        code = r[c_code] if c_code < len(r) else None
        if code is None:
            continue
        uw, qty = _f(_val(r, c_uw)), _f(_val(r, c_qty))
        dun = _f(_val(r, c_dun)) or 0.0
        if uw and qty:
            bundle_kg = uw * qty + dun     # actual timber mass + dunnage
        else:
            bundle_kg = _f(_val(r, c_kg_inc, c_kg))   # fallback: master's column
        skus[str(code).strip()] = {
            "qty_nop": qty,
            "unit_length_mm": _f(_val(r, c_len)),
            "bundle_height_mm": _f(_val(r, c_h_inc, c_h)),
            "bundle_width_mm": _f(_val(r, c_w)),
            "bundle_cubes_m3": _f(_val(r, c_m3_inc, c_m3)),
            "bundle_kg": bundle_kg,
        }

    # Read the orders tab by HEADER NAME, not fixed column positions, so
    # adding/removing/reordering columns in the tab never breaks the app.
    # Orders must come from the daily upload itself -- never fall back to
    # the reference file here, or the app would silently build yesterday's loads.
    ws = _find_sheet(wb, "order", use_reference=False)
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    header_idx, col = None, {}
    for i, r in enumerate(all_rows[:10]):
        names = [_norm(v) for v in r]
        if "due" in names and "sku" in names:
            header_idx = i
            col = {n: j for j, n in enumerate(names) if n}
            break
    if header_idx is None:
        raise ValueError('Could not find the header row on the "Fully Allocated Orders" tab '
                         '(expected a row containing "Due" and "SKU").')

    def _pick(row, *names):
        for n in names:
            if n in col:
                return row[col[n]]
        return None

    required = [("due",), ("sales order",), ("c/d",), ("location code",),
                ("delivery name",), ("site",), ("sku",), ("m3",), ("bundles",)]
    missing = [names[0] for names in required if not any(n in col for n in names)]
    if missing:
        raise ValueError('The "Fully Allocated Orders" tab is missing expected column(s): %s. '
                         'Found columns: %s' % (", ".join(missing), ", ".join(sorted(col))))

    orders = []
    excluded_collects = 0
    for r in all_rows[header_idx + 1:]:
        cd = _pick(r, "c/d")
        sku = _pick(r, "sku")
        if cd == "C":
            excluded_collects += 1
            continue
        if cd != "D" or sku is None:
            continue
        loc = _pick(r, "location code")
        site = _pick(r, "site")
        m3 = _pick(r, "m3")
        bundles = _pick(r, "bundles")
        m3_bundle = _pick(r, "m3 per bundle", "m3/bundle", "m3 bundle")
        if m3_bundle is None and m3 and bundles:
            m3_bundle = m3 / bundles
        orders.append({
            "due": _pick(r, "due"), "sales_order": _pick(r, "sales order"),
            "location_code": str(loc).strip() if loc is not None else "",
            "delivery_name": _pick(r, "delivery name"),
            "area": _pick(r, "area"), "prov": _pick(r, "prov", "province"),
            "site_raw": site, "site": _site_code(site),
            "sku": sku, "m3": m3, "m3_per_bundle": m3_bundle, "bundles": bundles,
        })

    _parse_demand_buckets(wb)

    return sites, customers, skus, orders, excluded_collects


def _order_digits(v):
    """SFP-000151714-SO -> 000151714 (digits core) for tolerant matching."""
    import re
    m = re.findall(r"\d+", str(v or ""))
    return max(m, key=len) if m else str(v or "").strip()


def _parse_demand_buckets(wb):
    """Best-effort parse of the Demand Buckets tab (upload first, then the
    reference file). Populates DEMAND_BUCKETS keyed by (order_digits, sku)
    and by sku alone as fallback. Never raises -- the Transport Orders tab
    simply gets blank bucket IDs if this tab is absent or unrecognisable."""
    global DEMAND_BUCKETS
    DEMAND_BUCKETS = {}
    ws = _find_sheet_or_none(wb, "demand") or _find_sheet_or_none(wb, "bucket")
    if ws is None:
        global _REF_WB
        if _REF_WB is None and os.path.exists(REFERENCE_FILE):
            _REF_WB = openpyxl.load_workbook(REFERENCE_FILE, data_only=True)
        if _REF_WB is not None:
            ws = _find_sheet_or_none(_REF_WB, "demand") or _find_sheet_or_none(_REF_WB, "bucket")
    if ws is None:
        return
    try:
        rows = list(ws.iter_rows(values_only=True))
        hdr_i, cm = None, {}
        for i, r in enumerate(rows[:12]):
            names = [_norm(v) for v in r]
            if any("sku" in n or "item" in n for n in names):
                hdr_i = i
                for j, n in enumerate(names):
                    if n and n not in cm:
                        cm[n] = j
                break
        if hdr_i is None:
            return

        def col(*words):
            for k, j in cm.items():
                if all(w in k for w in words):
                    return j
            return None

        c_sku = col("sku") if col("sku") is not None else col("item")
        c_lineno = col("line", "no") if col("line", "no") is not None else col("lineid")
        c_lineid = col("line", "id")
        c_bext = col("external") if col("external") is not None else col("bucket")
        c_ord = col("order")
        if c_sku is None:
            return
        # This tab lists one header row per bucket (External ID / Shipper
        # Reference / etc. filled in) followed by several blank-fronted
        # "Bucket Lines/..." continuation rows for that same bucket's SKUs.
        # There's no column literally named "order" in the real export --
        # the bucket's own External ID *is* the Sales Order text (e.g.
        # "SFP-000151617-SO" matches the orders tab exactly), so fall back
        # to it as the grouping key, carrying both it and the digits
        # forward across the blank continuation rows.
        last_order = None
        last_bext = ""
        for r in rows[hdr_i + 1:]:
            sku = r[c_sku] if c_sku < len(r) else None
            if sku is None:
                continue
            sku = str(sku).strip()
            ord_val = r[c_ord] if c_ord is not None and c_ord < len(r) else None
            if ord_val in (None, "") and c_ord is None and c_bext is not None and c_bext < len(r):
                ord_val = r[c_bext]
            if ord_val not in (None, ""):
                last_order = _order_digits(ord_val)
            bext_val = r[c_bext] if c_bext is not None and c_bext < len(r) else None
            if bext_val not in (None, ""):
                last_bext = bext_val
            rec = {
                "line_id": (r[c_lineid] if c_lineid is not None and c_lineid < len(r) else "") or "",
                "bucket_ext": last_bext,
                "line_no": (r[c_lineno] if c_lineno is not None and c_lineno < len(r) else "") or "",
            }
            if last_order:
                DEMAND_BUCKETS[(last_order, sku)] = rec
            DEMAND_BUCKETS.setdefault(sku, rec)
    except Exception:
        DEMAND_BUCKETS = {}


def bearing_degrees(lat1, lon1, lat2, lon2):
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dlambda = math.radians(lon2 - lon1)
    x = math.sin(dlambda) * math.cos(phi2)
    y = math.cos(phi1) * math.sin(phi2) - math.sin(phi1) * math.cos(phi2) * math.cos(dlambda)
    theta = math.atan2(x, y)
    return (math.degrees(theta) + 360) % 360


def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def _circular_angle_diff(a, b):
    d = abs(a - b) % 360
    return min(d, 360 - d)


def _circular_mean(angles):
    if not angles:
        return 0.0
    sx = sum(math.sin(math.radians(a)) for a in angles)
    cx = sum(math.cos(math.radians(a)) for a in angles)
    return math.degrees(math.atan2(sx, cx)) % 360


def enrich_lines(orders, customers, skus, sites):
    enriched = []
    for o in orders:
        cust = customers.get(o["location_code"])
        sku = skus.get(o["sku"])
        site = sites.get(o["site"])
        if cust is None or sku is None or site is None:
            continue
        dist_km = haversine_km(site["lat"], site["lon"], cust["lat"], cust["lon"])
        brg = bearing_degrees(site["lat"], site["lon"], cust["lat"], cust["lon"])
        line = dict(o)
        line.update({
            "cust_lat": cust["lat"], "cust_lon": cust["lon"], "province_full": cust["province"],
            "dist_km": dist_km, "bearing": brg,
            "dir_bucket": int(brg // DIRECTION_DEGREES),
            "bundle_length_m": (sku["unit_length_mm"] or 0) / 1000,
            "bundle_height_m": (sku["bundle_height_mm"] or 0) / 1000,
            "bundle_width_m": (sku["bundle_width_mm"] or 0) / 1000,
            "bundle_cubes_m3": sku["bundle_cubes_m3"] or 0,
            "bundle_kg": sku["bundle_kg"] or 0,
            "qty_nop": sku.get("qty_nop") or 0,
        })
        enriched.append(line)
    return enriched


def group_key(line):
    if GROUP_MODE == "province":
        return (line["site"], line["prov"])
    return (line["site"], line["dir_bucket"])


# ---------------------------------------------------------------------------
# Shared batch-filling + truck-selection helpers
# ---------------------------------------------------------------------------
def _greedy_fill_batch(candidates, max_drops):
    """Best-effort bin-covering: scan candidates in priority order, add
    whichever ones fit (skip -- don't stop -- on ones that don't), so a
    smaller order further down the list can still fill space a bigger one
    couldn't. Returns (batch, batch_m3, batch_kg, custset, leftover)."""
    batch, batch_m3, batch_kg, custset = [], 0.0, 0.0, set()
    leftover = []
    cap_m3 = TRUCK_TYPES["34T"]["cube_cap_m3"]
    cap_kg = TRUCK_TYPES["34T"]["payload_cap_t"] * 1000
    for ln in candidates:
        new_custset = custset | {ln["location_code"]}
        would_exceed_drops = len(new_custset) > max_drops
        would_exceed_cap = (batch_m3 + ln["m3"] > cap_m3) or (batch_kg + ln["bundle_kg"] * ln["bundles"] > cap_kg)
        if would_exceed_drops or would_exceed_cap:
            leftover.append(ln)
            continue
        batch.append(ln)
        batch_m3 += ln["m3"]
        batch_kg += ln["bundle_kg"] * ln["bundles"]
        custset = new_custset
    if not batch and candidates:
        ln = candidates[0]
        batch = [ln]
        batch_m3 = ln["m3"]
        batch_kg = ln["bundle_kg"] * ln["bundles"]
        custset = {ln["location_code"]}
        leftover = candidates[1:]
    return batch, batch_m3, batch_kg, custset, leftover


def _choose_truck_type(batch_m3, batch_kg, fleet_left, is_last_batch):
    batch_t = batch_kg / 1000
    for tname in TRUCK_TRY_ORDER:
        spec = TRUCK_TYPES[tname]
        if fleet_left[tname] <= 0:
            continue
        meets_min = (batch_t >= spec["payload_cap_t"] * MIN_WT_UTIL_PCT / 100.0
                     or batch_m3 >= spec["cube_cap_m3"] * MIN_VOL_UTIL_PCT / 100.0)
        fits = batch_t <= spec["payload_cap_t"] and batch_m3 <= spec["cube_cap_m3"]
        if fits and (meets_min or is_last_batch):
            return tname
    for tname in TRUCK_TRY_ORDER:
        spec = TRUCK_TYPES[tname]
        if fleet_left[tname] > 0 and batch_t <= spec["payload_cap_t"] and batch_m3 <= spec["cube_cap_m3"]:
            return tname
    return None


# ---------------------------------------------------------------------------
# Fixed-bucket assembly (direction / province toggle, as literally specified)
# ---------------------------------------------------------------------------
def _assemble_loads_fixed(lines):
    fleet_left = {k: v["fleet_count"] for k, v in TRUCK_TYPES.items()}
    groups = defaultdict(list)
    for ln in lines:
        groups[group_key(ln)].append(ln)

    loads = []
    unassigned = []
    load_id_counter = 1

    for gkey, glines in groups.items():
        glines.sort(key=lambda x: (x["due"] is None, x["due"]))
        pending = list(glines)

        while pending:
            batch, batch_m3, batch_kg, custset, leftover = _greedy_fill_batch(pending, MAX_DROPS_PER_LOAD)
            is_last_batch = len(leftover) == 0
            chosen_truck = _choose_truck_type(batch_m3, batch_kg, fleet_left, is_last_batch)

            if chosen_truck is None:
                unassigned.extend(batch)
                pending = leftover
                continue

            fleet_left[chosen_truck] -= 1
            loads.append({
                "load_id": f"L{load_id_counter:03d}",
                "site": gkey[0],
                "group": gkey[1],
                "truck_type": chosen_truck,
                "lines": batch,
                "total_m3": batch_m3,
                "total_kg": batch_kg,
                "n_customers": len(custset),
            })
            load_id_counter += 1
            pending = leftover

    return loads, unassigned, fleet_left


# ---------------------------------------------------------------------------
# Adaptive assembly: seed each load with the oldest-due pending order, then
# pull in the nearest-bearing other pending orders (regardless of a fixed
# bucket boundary) until the truck is as full as it can be.
# ---------------------------------------------------------------------------
def _assemble_loads_adaptive(lines):
    fleet_left = {k: v["fleet_count"] for k, v in TRUCK_TYPES.items()}
    by_site = defaultdict(list)
    for ln in lines:
        by_site[ln["site"]].append(ln)

    loads = []
    unassigned = []
    load_id_counter = 1

    for site, site_lines in by_site.items():
        pending = list(site_lines)

        while pending:
            pending.sort(key=lambda x: (x["due"] is None, x["due"]))
            seed_bearing = pending[0]["bearing"]
            # Hard corridor cap: only ever consider freight within
            # MAX_CORRIDOR_SPREAD_DEG of the seed's bearing. Freight outside
            # that spread is left in "pending" for a later batch/load --
            # never pulled in just to fill out this truck.
            in_corridor = [x for x in pending
                           if _circular_angle_diff(x["bearing"], seed_bearing) <= MAX_CORRIDOR_SPREAD_DEG]
            candidates = sorted(in_corridor, key=lambda x: _circular_angle_diff(x["bearing"], seed_bearing))

            batch, batch_m3, batch_kg, custset, leftover = _greedy_fill_batch(candidates, MAX_DROPS_PER_LOAD)
            batch_ids = {id(x) for x in batch}
            new_pending = [x for x in pending if id(x) not in batch_ids]
            is_last_batch = len(new_pending) == 0

            chosen_truck = _choose_truck_type(batch_m3, batch_kg, fleet_left, is_last_batch)

            if chosen_truck is None:
                unassigned.extend(batch)
                pending = new_pending
                continue

            fleet_left[chosen_truck] -= 1
            bearings = [ln["bearing"] for ln in batch]
            centroid = _circular_mean(bearings)
            spread = max((_circular_angle_diff(b, centroid) for b in bearings), default=0)
            loads.append({
                "load_id": f"L{load_id_counter:03d}",
                "site": site,
                "group": ("adaptive", centroid, spread),
                "truck_type": chosen_truck,
                "lines": batch,
                "total_m3": batch_m3,
                "total_kg": batch_kg,
                "n_customers": len(custset),
            })
            load_id_counter += 1
            pending = new_pending

    return loads, unassigned, fleet_left


def _split_line(ln, n_evict):
    """Split n_evict bundles off a line. Returns (kept_or_None, evicted)."""
    per_b_m3 = ln["m3"] / ln["bundles"] if ln["bundles"] else 0.0
    evicted = dict(ln)
    evicted["bundles"] = n_evict
    evicted["m3"] = per_b_m3 * n_evict
    kept = None
    if n_evict < ln["bundles"]:
        kept = dict(ln)
        kept["bundles"] = ln["bundles"] - n_evict
        kept["m3"] = per_b_m3 * kept["bundles"]
    return kept, evicted


def _line_fits_load_group(ln, load):
    """Would this line be an acceptable match for an ALREADY-BUILT load,
    under the same grouping rule that built it -- corridor spread for
    adaptive mode, the same fixed bucket for direction/province modes?
    Used only to decide whether evicted (physically-didn't-fit) freight
    may be offered a spare seat on an existing load, so it never crosses
    a corridor/province boundary just to avoid a near-empty extra truck."""
    if load["site"] != ln["site"]:
        return False
    g = load["group"]
    if isinstance(g, tuple) and g and g[0] == "adaptive":
        _, centroid, _spread = g
        return _circular_angle_diff(ln["bearing"], centroid) <= MAX_CORRIDOR_SPREAD_DEG
    if GROUP_MODE == "province":
        return g == ln["prov"]
    return g == ln["dir_bucket"]


def _try_salvage_into_existing(evicted_lines, kept_loads):
    """Before spinning up a dedicated (often near-empty) new truck for
    freight that didn't physically fit its original load (floor/stack/
    height limits, NOT capacity -- capacity overflow is already handled
    by folding back into the next batch during initial assembly), first
    see if it can slot into spare room on an ALREADY-BUILT load that is
    (a) from the same site, (b) still an acceptable match under the same
    corridor/direction/province rule that built it, and (c) has spare
    weight/volume/drop-count headroom. Only what genuinely can't be
    salvaged this way falls through to a fresh assemble() round that may
    need to build new trucks."""
    still_evicted = []
    for ln in evicted_lines:
        placed = False
        for ld in kept_loads:
            if not _line_fits_load_group(ln, ld):
                continue
            spec = TRUCK_TYPES[ld["truck_type"]]
            added_kg = ln["bundle_kg"] * ln["bundles"]
            added_m3 = ln["m3"]
            if ld["total_kg"] + added_kg > spec["payload_cap_t"] * 1000 + 1e-6:
                continue
            if ld["total_m3"] + added_m3 > spec["cube_cap_m3"] + 1e-6:
                continue
            cust_set = {l["location_code"] for l in ld["lines"]}
            if ln["location_code"] not in cust_set and len(cust_set) >= MAX_DROPS_PER_LOAD:
                continue

            trial_lines = ld["lines"] + [ln]
            trial_load = dict(ld)
            trial_load["lines"] = trial_lines
            packing, leftover = pack_load(trial_load)
            if leftover:
                continue  # doesn't actually fit once re-simulated -- try the next load

            ld["lines"] = trial_lines
            ld["total_m3"] += added_m3
            ld["total_kg"] += added_kg
            ld["n_customers"] = len({l["location_code"] for l in trial_lines})
            ld["packing"] = packing
            ld["pack_leftover"] = leftover
            placed = True
            break
        if not placed:
            still_evicted.append(ln)
    return still_evicted


def resize_load_for_truck(load, new_truck_key, unassigned_pool):
    """Re-plan a single already-built load for a NEW truck type -- used by
    the 'truck type override' Rebuild action, so switching a load to a
    bigger/smaller truck actually changes what's ON it, not just whether the
    same freight still physically fits.

    Step 1 (shrink): if the load's current freight no longer fits the new
    truck's weight/cube caps, evict the least-urgent lines (latest due date
    first) back into `unassigned_pool` until it fits.
    Step 2 (grow): if the new truck has spare capacity, pull in nearby
    freight sitting in `unassigned_pool` -- same site, same corridor/
    direction/province this load was originally built under (via
    `_line_fits_load_group`) -- oldest-due first, same as a fresh build
    would for this truck size.

    Mutates `load` (truck_type, lines, total_m3, total_kg, n_customers) and
    `unassigned_pool` (evicted lines appended, consumed lines removed) in
    place. Caller should re-run pack_load() afterwards to get the real
    placement/leftover for the new truck.
    """
    max_drops = MAX_DROPS_PER_LOAD
    spec = TRUCK_TYPES[new_truck_key]
    cap_kg = spec["payload_cap_t"] * 1000
    cap_m3 = spec["cube_cap_m3"]

    load["truck_type"] = new_truck_key

    lines = sorted(load["lines"], key=lambda x: (x["due"] is None, x["due"]))
    kept, evicted = [], []
    kg = m3 = 0.0
    custset = set()
    for ln in lines:
        new_custset = custset | {ln["location_code"]}
        add_kg = ln["bundle_kg"] * ln["bundles"]
        add_m3 = ln["m3"]
        would_exceed = (kg + add_kg > cap_kg + 1e-6 or m3 + add_m3 > cap_m3 + 1e-6
                        or len(new_custset) > max_drops)
        if would_exceed:
            evicted.append(ln)
            continue
        kept.append(ln)
        kg += add_kg
        m3 += add_m3
        custset = new_custset
    if evicted:
        unassigned_pool.extend(evicted)

    # Spare capacity on the new truck -- top it up from nearby unassigned
    # freight, same as a fresh build would.
    candidates = [ln for ln in unassigned_pool if _line_fits_load_group(ln, load)]
    candidates.sort(key=lambda x: (x["due"] is None, x["due"]))
    consumed = []
    for ln in candidates:
        new_custset = custset | {ln["location_code"]}
        add_kg = ln["bundle_kg"] * ln["bundles"]
        add_m3 = ln["m3"]
        if kg + add_kg > cap_kg + 1e-6 or m3 + add_m3 > cap_m3 + 1e-6 or len(new_custset) > max_drops:
            continue
        kept.append(ln)
        kg += add_kg
        m3 += add_m3
        custset = new_custset
        consumed.append(ln)
    for ln in consumed:
        unassigned_pool.remove(ln)

    load["lines"] = kept
    load["total_kg"] = kg
    load["total_m3"] = m3
    load["n_customers"] = len(custset)
    return load


def assemble_loads(lines):
    """Assemble loads on paper, then RECONCILE each with the physical
    packer: any bundle that doesn't physically fit (height/floor/stack
    limits) is evicted from that load and rolled into a fresh assembly
    round on the remaining fleet -- so a load's paperwork always matches
    what actually fits on the truck, and overflow freight gets its own
    truck instead of silently falling off the plan."""
    assemble = _assemble_loads_adaptive if GROUP_MODE == "adaptive" else _assemble_loads_fixed
    loads, unassigned, fleet_left = assemble(lines)

    all_loads = []
    for round_no in range(6):
        evicted_lines = []
        kept_loads = []
        for load in loads:
            packing, leftover = pack_load(load)
            if leftover:
                evict_n = defaultdict(int)
                for u in leftover:
                    evict_n[(u["sales_order"], u["sku"], u["location_code"])] += 1
                new_lines = []
                for ln in load["lines"]:
                    k = (ln["sales_order"], ln["sku"], ln["location_code"])
                    n = min(evict_n.get(k, 0), ln["bundles"])
                    if n:
                        evict_n[k] -= n
                        kept, ev = _split_line(ln, n)
                        evicted_lines.append(ev)
                        if kept:
                            new_lines.append(kept)
                    else:
                        new_lines.append(ln)
                if not new_lines:
                    # nothing fit at all -- release the truck, unassign freight
                    fleet_left[load["truck_type"]] += 1
                    unassigned.extend(evicted_lines[-1:])
                    continue
                load["lines"] = new_lines
                load["total_m3"] = sum(l["m3"] for l in new_lines)
                load["total_kg"] = sum(l["bundle_kg"] * l["bundles"] for l in new_lines)
                load["n_customers"] = len({l["location_code"] for l in new_lines})
            kept_loads.append(load)

        if evicted_lines:
            evicted_lines = _try_salvage_into_existing(evicted_lines, kept_loads)

        all_loads.extend(kept_loads)

        if not evicted_lines:
            break
        if round_no == 5:
            unassigned.extend(evicted_lines)
            break

        # assemble the evicted freight onto whatever fleet remains
        saved = {k: TRUCK_TYPES[k]["fleet_count"] for k in TRUCK_TYPES}
        for k in TRUCK_TYPES:
            TRUCK_TYPES[k]["fleet_count"] = fleet_left[k]
        try:
            loads, more_unassigned, fleet_left = assemble(evicted_lines)
        finally:
            for k in TRUCK_TYPES:
                TRUCK_TYPES[k]["fleet_count"] = saved[k]
        unassigned.extend(more_unassigned)
        if not loads:
            break

    for i, load in enumerate(all_loads, 1):
        load["load_id"] = "L%03d" % i
    return all_loads, unassigned, fleet_left


OVERHANG_ALLOW = 0.15


def _new_trailer_state(trailer_spec):
    length_cap = trailer_spec["length_m"]
    n_slots = int(trailer_spec.get("width_slots", 2))
    return {
        "skyline": [[{"x0": 0.0, "x1": length_cap, "h": 0.0}] for _ in range(n_slots)],
        "used_weight": 0.0, "used_volume": 0.0, "placements": [],
    }


LEVEL_TOL = 0.05   # adjacent stacks within 5 cm of the same height count as
                   # one level surface -- loaders bridge these with dunnage
                   # bearers, letting a long bundle lie across two stacks.


def _try_place_unit(state, trailer_spec, u, allowed_slots=None):
    """Best-fit skyline placement. A bundle may rest within one shelf
    segment OR span a run of consecutive near-level segments (real loading
    practice: a 6m bundle lies across two level 3m stacks). Leftover shelf
    length beside a shorter bundle stays open for later bundles.
    `allowed_slots` restricts placement to specific lane index(es) -- used
    by the interactive editor to pin a bundle to a chosen lane."""
    height_cap = trailer_spec["height_m"]
    weight_cap = trailer_spec["weight_cap_t"] * 1000
    if state["used_weight"] + u["bundle_kg"] > weight_cap:
        return False

    best = None
    slot_range = range(len(state["skyline"])) if allowed_slots is None else \
        [s for s in allowed_slots if 0 <= s < len(state["skyline"])]
    for slot_idx in slot_range:
        slot = state["skyline"][slot_idx]
        for i in range(len(slot)):
            if slot[i]["x1"] - slot[i]["x0"] <= 1e-9:
                continue
            base_h = slot[i]["h"]
            j = i
            while True:
                span_len = slot[j]["x1"] - slot[i]["x0"]
                h_top = max(slot[k]["h"] for k in range(i, j + 1))
                on_floor = h_top <= 1e-9
                max_allowed = span_len if on_floor else span_len * (1 + OVERHANG_ALLOW)
                if (u["bundle_length_m"] <= max_allowed + 1e-9
                        and h_top + u["bundle_height_m"] <= height_cap + 1e-9):
                    overhang = max(0.0, u["bundle_length_m"] - span_len)
                    waste = abs(span_len - u["bundle_length_m"])
                    score = (1 if overhang > 1e-9 else 0,
                             round(h_top + u["bundle_height_m"], 4), waste, j - i)
                    if best is None or score < best[0]:
                        best = (score, slot_idx, i, j, h_top)
                    break  # fits on this run; extending further only adds waste
                # too short so far -- extend across the next segment if its
                # top is level (within dunnage tolerance) with the run's base
                if j + 1 < len(slot) and abs(slot[j + 1]["h"] - base_h) <= LEVEL_TOL:
                    j += 1
                else:
                    break

    if best is None:
        return False

    _, slot_idx, i, j, h_top = best
    slot = state["skyline"][slot_idx]
    x0 = slot[i]["x0"]
    span_len = slot[j]["x1"] - x0
    claim_len = min(u["bundle_length_m"], span_len)
    claim_x1 = x0 + claim_len

    new_segments = [{"x0": x0, "x1": claim_x1, "h": h_top + u["bundle_height_m"]}]
    last = slot[j]
    if claim_x1 < last["x1"] - 1e-9:
        new_segments.append({"x0": claim_x1, "x1": last["x1"], "h": last["h"]})
    slot[i:j + 1] = new_segments

    # merge adjacent equal-height segments so later bundles see one long
    # continuous shelf instead of fragmented slivers (e.g. two 4.8m bundles
    # side by side leave a single 2.4m opening, not scattered scraps)
    merged = [slot[0]]
    for seg in slot[1:]:
        if abs(seg["h"] - merged[-1]["h"]) <= 1e-6:
            merged[-1] = {"x0": merged[-1]["x0"], "x1": seg["x1"], "h": seg["h"]}
        else:
            merged.append(seg)
    slot[:] = merged

    state["used_weight"] += u["bundle_kg"]
    state["used_volume"] += u.get("bundle_cubes_m3", 0)
    state["placements"].append({"slot": slot_idx, "x": x0, "y": h_top, **u})
    return True


def _used_length(state):
    """Average occupied floor length across BOTH width-slots (not just the
    furthest reach of whichever slot has the most freight in it) -- a
    trailer with one lane full and the other lane empty is genuinely only
    ~50% floor-utilised, not 100%, since the floor is 2 lanes wide."""
    total = 0.0
    for slot in state["skyline"]:
        total += sum(seg["x1"] - seg["x0"] for seg in slot if seg["h"] > 1e-9)
    return total / len(state["skyline"])


def _expand_units(bundles):
    units = []
    for b in bundles:
        for _ in range(int(b["bundles"])):
            units.append(dict(b))
    units.sort(key=lambda u: (-u["bundle_length_m"], -u["bundle_kg"]))
    # Stable unit id: k-th identical bundle of (order, sku, customer).
    # Identical units are physically interchangeable, so pinning "the k-th
    # unit of kind X" stays meaningful across repacks even if the load's
    # line set grows or shrinks.
    counts = defaultdict(int)
    for u in units:
        k = (u["sales_order"], u["sku"], u["location_code"])
        u["uid"] = "%s|%s|%s#%d" % (k[0], k[1], k[2], counts[k])
        counts[k] += 1
    return units


def _order_units(units, unit_sort):
    """Re-order units for a packing attempt. Default (None/'length') keeps
    the longest-first order; other strategies give the optimizer different
    starting sequences to search over."""
    if not unit_sort or unit_sort == "length":
        return units
    if unit_sort == "height":
        units.sort(key=lambda u: (-u["bundle_height_m"], -u["bundle_length_m"], -u["bundle_kg"]))
    elif unit_sort == "weight":
        units.sort(key=lambda u: (-u["bundle_kg"], -u["bundle_length_m"]))
    elif unit_sort == "volume":
        units.sort(key=lambda u: (-u.get("bundle_cubes_m3", 0), -u["bundle_length_m"]))
    elif unit_sort.startswith("seed"):
        import random
        random.Random(int(unit_sort[4:])).shuffle(units)
    return units


def pack_trailer(trailer_spec, bundles, pins=None, unit_sort=None):
    """Single-trailer packer (kept for compatibility / single-trailer trucks).
    `pins` maps unit uid -> (trailer_name, slot_idx or None): pinned units are
    placed first, restricted to their chosen lane, then the rest re-flow
    around them."""
    units = _order_units(_expand_units(bundles), unit_sort)
    state = _new_trailer_state(trailer_spec)
    pins = pins or {}
    leftover = []
    pinned = [u for u in units if u["uid"] in pins]
    rest = [u for u in units if u["uid"] not in pins]
    for u in pinned:
        _, slot = pins[u["uid"]]
        if not _try_place_unit(state, trailer_spec, u,
                               allowed_slots=None if slot is None else [slot]):
            rest.append(u)  # pin impossible -- fall back to normal placement
    for u in rest:
        if not _try_place_unit(state, trailer_spec, u):
            leftover.append(u)
    return state["placements"], leftover, _used_length(state), state["used_weight"], state["used_volume"]


def _assign_bids(packing):
    """Sequential display numbers (b1, b2, ...) across a load's trailers so
    the on-screen editor and schematic can reference individual bundles."""
    bid = 1
    for name in packing:
        for p in sorted(packing[name]["placements"], key=lambda p: (p["slot"], p["x"], p["y"])):
            p["bid"] = bid
            bid += 1
    return packing


def pack_load(load, pins=None, unit_sort=None):
    spec = TRUCK_TYPES[load["truck_type"]]
    trailers = spec["trailers"]
    total_trailer_length = sum(t["length_m"] for t in trailers)

    bundle_recs = []
    for ln in load["lines"]:
        bundle_recs.append({
            "bundle_length_m": ln["bundle_length_m"], "bundle_height_m": ln["bundle_height_m"],
            "bundle_width_m": ln["bundle_width_m"], "bundle_kg": ln["bundle_kg"],
            "bundle_cubes_m3": ln["bundle_cubes_m3"],
            "bundles": ln["bundles"], "sales_order": ln["sales_order"], "sku": ln["sku"],
            "delivery_name": ln["delivery_name"], "location_code": ln["location_code"],
            "dist_km": ln["dist_km"],
        })

    if len(trailers) == 1:
        placements, leftover, used_len, used_wt, used_vol = pack_trailer(
            trailers[0], bundle_recs, pins=pins, unit_sort=unit_sort)
        packing = {trailers[0]["name"]: {
            "placements": placements, "spec": trailers[0],
            "used_length": used_len, "used_weight": used_wt, "used_volume": used_vol,
            "cube_cap_m3": spec["cube_cap_m3"],
        }}
        return _assign_bids(packing), leftover

    front_spec = next(t for t in trailers if t["name"] == "front")
    rear_spec = next(t for t in trailers if t["name"] == "rear")
    front_state = _new_trailer_state(front_spec)
    rear_state = _new_trailer_state(rear_spec)

    units = _order_units(_expand_units(bundle_recs), unit_sort)
    pins = pins or {}
    trailer_states = {"front": (front_state, front_spec), "rear": (rear_state, rear_spec)}

    # Pinned units first: each goes into its chosen trailer (and lane, if
    # given); the automatic balanced flow below then re-flows everything
    # else around them. A pin that physically can't be honoured falls back
    # to normal placement rather than dropping the bundle.
    pinned_units = [u for u in units if u["uid"] in pins]
    units = [u for u in units if u["uid"] not in pins]
    failed_pins = []
    for u in pinned_units:
        tname, slot = pins[u["uid"]]
        st_sp = trailer_states.get(tname)
        ok = False
        if st_sp is not None:
            ok = _try_place_unit(st_sp[0], st_sp[1], u,
                                 allowed_slots=None if slot is None else [slot])
        if not ok:
            failed_pins.append(u)
    units = failed_pins + units

    # Balanced two-trailer allocation: give every DROP a presence in BOTH
    # trailers, in proportion to each trailer's weight capacity (1/3 front,
    # 2/3 rear on a 34T interlink). As the route is unloaded drop by drop,
    # both trailers then empty gradually together -- avoiding the risk case
    # of a nearly-empty front pulling a still-full 12m rear trailer.
    front_share = front_spec["weight_cap_t"] / (front_spec["weight_cap_t"] + rear_spec["weight_cap_t"])
    drop_front_kg = defaultdict(float)
    drop_total_kg = defaultdict(float)
    for p in front_state["placements"]:
        drop_front_kg[p["location_code"]] += p["bundle_kg"]
        drop_total_kg[p["location_code"]] += p["bundle_kg"]
    for p in rear_state["placements"]:
        drop_total_kg[p["location_code"]] += p["bundle_kg"]

    leftover = []
    for u in units:
        d = u["location_code"]
        want_front = drop_front_kg[d] < front_share * (drop_total_kg[d] + u["bundle_kg"])
        order = ((front_state, front_spec), (rear_state, rear_spec)) if want_front \
            else ((rear_state, rear_spec), (front_state, front_spec))
        placed_state = None
        for st, sp in order:
            if _try_place_unit(st, sp, u):
                placed_state = st
                break
        if placed_state is None:
            leftover.append(u)
            continue
        drop_total_kg[d] += u["bundle_kg"]
        if placed_state is front_state:
            drop_front_kg[d] += u["bundle_kg"]

    # Second-chance pass: placements made after a bundle first failed can
    # create new level surfaces it now fits on -- try leftovers once more.
    still_left = []
    for u in leftover:
        if _try_place_unit(rear_state, rear_spec, u) or _try_place_unit(front_state, front_spec, u):
            continue
        still_left.append(u)
    leftover = still_left

    front_cube_cap = spec["cube_cap_m3"] * front_spec["length_m"] / total_trailer_length
    rear_cube_cap = spec["cube_cap_m3"] * rear_spec["length_m"] / total_trailer_length

    packing = {
        "front": {"placements": front_state["placements"], "spec": front_spec,
                  "used_length": _used_length(front_state), "used_weight": front_state["used_weight"],
                  "used_volume": front_state["used_volume"], "cube_cap_m3": front_cube_cap},
        "rear": {"placements": rear_state["placements"], "spec": rear_spec,
                 "used_length": _used_length(rear_state), "used_weight": rear_state["used_weight"],
                 "used_volume": rear_state["used_volume"], "cube_cap_m3": rear_cube_cap},
    }
    return _assign_bids(packing), leftover


def drop_balance_report(load):
    """Warnings (not errors) for drops whose weight sits much more heavily
    in one trailer than the balanced front/rear share -- the risk case is a
    nearly-empty front trailer pulling a still-full rear as the route is
    unloaded drop by drop. Only meaningful for two-trailer trucks."""
    packing = load.get("packing") or {}
    if "front" not in packing or "rear" not in packing:
        return []
    f_spec = packing["front"]["spec"]
    r_spec = packing["rear"]["spec"]
    share = f_spec["weight_cap_t"] / (f_spec["weight_cap_t"] + r_spec["weight_cap_t"])
    front_kg = defaultdict(float)
    total_kg = defaultdict(float)
    names = {}
    for p in packing["front"]["placements"]:
        front_kg[p["location_code"]] += p["bundle_kg"]
        total_kg[p["location_code"]] += p["bundle_kg"]
        names[p["location_code"]] = p.get("delivery_name") or p["location_code"]
    for p in packing["rear"]["placements"]:
        total_kg[p["location_code"]] += p["bundle_kg"]
        names.setdefault(p["location_code"], p.get("delivery_name") or p["location_code"])
    warnings = []
    for code, tot in total_kg.items():
        if tot < 500:
            continue  # too light to matter for balance
        frac = front_kg[code] / tot
        if abs(frac - share) > 0.35:
            where = "front" if frac > share else "rear"
            warnings.append(
                "Drop balance: %s has %.0f%% of its %.0f kg in the %s trailer "
                "(balanced would be ~%.0f%% front) -- fine if intentional, but the %s "
                "trailer will empty much faster at that stop." % (
                    names[code], frac * 100, tot, "front" if frac > share else "rear",
                    share * 100, where))
    return warnings


def refit_from_pool(load, unassigned_pool, pins=None, unit_sort=None):
    """Try to fill spare space on an already-packed load with freight from
    the unassigned pool -- same site + corridor/direction/province rule the
    load was built under, oldest due first. A candidate line is kept only
    if the physical repack places EVERY one of its bundles without bumping
    anything already placed. Mutates load and pool; returns list of lines
    that were added."""
    spec = TRUCK_TYPES[load["truck_type"]]
    cap_kg = spec["payload_cap_t"] * 1000
    cap_m3 = spec["cube_cap_m3"]
    baseline_left = len(load.get("pack_leftover") or [])
    added = []
    candidates = [ln for ln in unassigned_pool if _line_fits_load_group(ln, load)]
    candidates.sort(key=lambda x: (x["due"] is None, x["due"]))
    for ln in candidates:
        add_kg = ln["bundle_kg"] * ln["bundles"]
        if load["total_kg"] + add_kg > cap_kg + 1e-6 or load["total_m3"] + ln["m3"] > cap_m3 + 1e-6:
            continue
        cust = {l["location_code"] for l in load["lines"]}
        if ln["location_code"] not in cust and len(cust) >= MAX_DROPS_PER_LOAD:
            continue
        trial = dict(load)
        trial["lines"] = load["lines"] + [ln]
        packing, leftover = pack_load(trial, pins=pins, unit_sort=unit_sort)
        if len(leftover) > baseline_left:
            continue  # didn't physically fit (or bumped something) -- skip
        load["lines"] = trial["lines"]
        load["total_kg"] += add_kg
        load["total_m3"] += ln["m3"]
        load["n_customers"] = len({l["location_code"] for l in load["lines"]})
        load["packing"] = packing
        load["pack_leftover"] = leftover
        baseline_left = len(leftover)
        unassigned_pool.remove(ln)
        added.append(ln)
    return added


PACK_STRATEGIES = ["length", "height", "weight", "volume"] + ["seed%d" % i for i in range(1, 25)]


def optimize_load_packing(load, pins=None):
    """Search over packing orders (a few deterministic heuristics plus
    seeded shuffles) for the arrangement that places the most bundles --
    ties broken by most volume on board, then best floor use. Honours any
    pinned bundles. Updates the load in place; returns (strategy, packing
    improved: bool)."""
    def score(packing, leftover):
        placed_vol = sum(t["used_volume"] for t in packing.values())
        floor = sum(t["used_length"] for t in packing.values())
        return (len(leftover), -placed_vol, -floor)

    best = None
    for strat in PACK_STRATEGIES:
        packing, leftover = pack_load(load, pins=pins, unit_sort=strat)
        s = score(packing, leftover)
        if best is None or s < best[0]:
            best = (s, strat, packing, leftover)
    cur_left = len(load.get("pack_leftover") or [])
    cur_vol = sum(t["used_volume"] for t in (load.get("packing") or {}).values())
    _, strat, packing, leftover = best
    improved = (len(leftover) < cur_left) or (
        len(leftover) == cur_left and sum(t["used_volume"] for t in packing.values()) > cur_vol + 1e-9)
    load["packing"] = packing
    load["pack_leftover"] = leftover
    load["pack_strategy"] = strat
    return strat, improved


if __name__ == "__main__":
    sites, customers, skus, orders, excluded = load_workbook_data()
    print(f"Sites: {len(sites)}  Customers: {len(customers)}  SKUs: {len(skus)}")
    print(f"Delivery lines loaded: {len(orders)}  (collects excluded: {excluded})")
    lines = enrich_lines(orders, customers, skus, sites)
    print(f"Enriched lines (joined ok): {len(lines)}")

    loads, unassigned, fleet_left = assemble_loads(lines)
    print(f"\nMode: {GROUP_MODE}")
    print(f"Loads built: {len(loads)}")
    print(f"Unassigned lines: {len(unassigned)}")
    print(f"Fleet remaining: {fleet_left}")

    total_leftover = 0
    for load in loads:
        packing, leftover = pack_load(load)
        load["packing"] = packing
        load["pack_leftover"] = leftover
        total_leftover += len(leftover)
        util_m3 = load["total_m3"] / TRUCK_TYPES[load["truck_type"]]["cube_cap_m3"] * 100
        util_kg = load["total_kg"] / (TRUCK_TYPES[load["truck_type"]]["payload_cap_t"] * 1000) * 100
        print("  %s site=%s group=%s truck=%s m3=%.1f (%.0f%%) kg=%.0f (%.0f%%) drops=%d not_placed=%d"
              % (load["load_id"], load["site"], group_label(load["group"]), load["truck_type"],
                 load["total_m3"], util_m3, load["total_kg"], util_kg,
                 load["n_customers"], len(leftover)))

    print("Total bundles not physically placed: %d" % total_leftover)
