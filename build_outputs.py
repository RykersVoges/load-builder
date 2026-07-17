"""
Generates the two output Excel tabs (Loads Summary, Orders Line Summary)
and one landscape schematic page per load (multi-page PDF), from the
load_builder prototype's results.

Schematic v5: numeric length-axis tick labels along the bottom of each
trailer, explicit "Ord"/"SKU" tags so the two numbers on a bundle are never
ambiguous, a white halo behind every label so hatch lines never run through
the text, and lighter (single-pass) hatch patterns for less ink and a
cleaner look. Legend stays at the top in route order (closest drop first);
gap-filled skyline packing from the previous round is unchanged.
"""
from collections import defaultdict
import openpyxl

APP_VERSION = "v33 (17 Jul 2026)"
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.backends.backend_pdf import PdfPages

from load_builder import (
    load_workbook_data, enrich_lines, assemble_loads, pack_load, TRUCK_TYPES, group_label
)

NAVY = "#1F3352"
STEEL = "#4A6B8A"
INK = "#1A1A1A"

# Distinct customer styles used as thin borders/text accents + a light hatch
# pattern (never a solid fill) so the whole schematic is cheap to print in
# black ink. Single-character hatch codes = sparser lines than repeated ones.
ACCENTS = ["#4C78A8", "#F58518", "#54A24B", "#B2323C", "#4FA9A5",
           "#B8860B", "#7B4F9D", "#C2568B", "#8C5A2B", "#5C5C5C"]
HATCHES = ["/", "\\", "|", "-", "+", "x", ".", "o", "*", "\\|"]


def _drop_sequence(load):
    """Order drops by distance from site (closest first == drop 1 / front
    trailer, matching the loading/unloading sequence), returns
    {location_code: (seq_no, name, min_dist_km)}."""
    best = {}
    for t in load["packing"].values():
        for p in t["placements"]:
            code = p.get("location_code")
            d = p.get("dist_km", 0)
            if code not in best or d < best[code][1]:
                best[code] = (p.get("delivery_name", ""), d)
    ordered = sorted(best.items(), key=lambda kv: kv[1][1])
    return {code: (i + 1, name, dist) for i, (code, (name, dist)) in enumerate(ordered)}


def _customer_style_map(load):
    seq = _drop_sequence(load)
    codes = sorted(seq.keys(), key=lambda c: seq[c][0])
    return {code: (ACCENTS[i % len(ACCENTS)], HATCHES[i % len(HATCHES)]) for i, code in enumerate(codes)}

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)


def build_all():
    sites, customers, skus, orders, excluded = load_workbook_data()
    lines = enrich_lines(orders, customers, skus, sites)
    loads, unassigned, fleet_left = assemble_loads(lines)
    for load in loads:
        packing, leftover = pack_load(load)
        load["packing"] = packing
        load["pack_leftover"] = leftover
    return loads, unassigned, fleet_left, lines, sites


def sheet_style_print(ws, n_cols):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_options.horizontalCentered = True
    for i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16


def write_loads_summary(wb, loads, unassigned):
    ws = wb.active
    ws.title = "Loads Summary"
    headers = ["Load ID", "Site", "Truck Type", "Direction/Area Group",
               "Total m3", "Vol Util %", "Total KG", "Weight Util %",
               "# Orders", "Full/Partial", "# Unique SKUs", "# Customers/Drops",
               "# Bundles Placed", "Bundles Not Placed (packing)"]
    ws.append(headers)
    for c in ws[1]:
        c.font = BOLD
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    order_loads = defaultdict(set)
    for load in loads:
        for ln in load["lines"]:
            order_loads[ln["sales_order"]].add(load["load_id"])
    order_unassigned = defaultdict(bool)
    for ln in unassigned:
        order_unassigned[ln["sales_order"]] = True

    for load in loads:
        spec = TRUCK_TYPES[load["truck_type"]]
        vol_util = load["total_m3"] / spec["cube_cap_m3"] * 100
        wt_util = load["total_kg"] / (spec["payload_cap_t"] * 1000) * 100
        so_set = {ln["sales_order"] for ln in load["lines"]}
        sku_set = {ln["sku"] for ln in load["lines"]}
        cust_set = {ln["location_code"] for ln in load["lines"]}
        full_flags = []
        for so in so_set:
            is_full = len(order_loads[so]) == 1 and not order_unassigned[so]
            full_flags.append("Full" if is_full else "Partial")
        full_partial = "Full" if all(f == "Full" for f in full_flags) else "Partial"
        leftover_n = len(load.get("pack_leftover", []))
        total_bundles = sum(ln["bundles"] for ln in load["lines"])
        row = [load["load_id"], load["site"], load["truck_type"], group_label(load["group"]),
               round(load["total_m3"], 2), round(vol_util, 0), round(load["total_kg"], 0), round(wt_util, 0),
               len(so_set), full_partial, len(sku_set), len(cust_set),
               total_bundles - leftover_n, leftover_n]
        ws.append(row)
        for c in ws[ws.max_row]:
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")

    sheet_style_print(ws, len(headers))
    ws.freeze_panes = "A2"
    return ws


def write_orders_summary(wb, loads, unassigned):
    ws = wb.create_sheet("Orders Line Summary")
    headers = ["Sales Order", "Due Date", "Customer", "Location Code", "SKU",
               "Bundles", "Bundles Placed", "m3", "KG", "Assigned Load ID", "Site", "Status"]
    ws.append(headers)
    for c in ws[1]:
        c.font = BOLD
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    # How many bundles of each line did NOT physically fit in the packing
    # simulation -- so the Status column tells the truth per line, matching
    # the schematic's "x/y bundles placed" header exactly.
    not_placed = defaultdict(int)
    for load in loads:
        for u in load.get("pack_leftover", []):
            not_placed[(load["load_id"], u["sales_order"], u["sku"], u["location_code"])] += 1

    def add_row(ln, load_id, status, placed=None):
        kg = ln["bundle_kg"] * ln["bundles"]
        ws.append([ln["sales_order"], ln["due"].strftime("%Y-%m-%d") if ln["due"] else "",
                   ln["delivery_name"], ln["location_code"], ln["sku"], ln["bundles"],
                   placed if placed is not None else 0,
                   round(ln["m3"], 2), round(kg, 1), load_id, ln["site"], status])
        for c in ws[ws.max_row]:
            c.border = BORDER

    for load in loads:
        for ln in load["lines"]:
            miss = not_placed.get((load["load_id"], ln["sales_order"], ln["sku"], ln["location_code"]), 0)
            placed = ln["bundles"] - miss
            if miss == 0:
                add_row(ln, load["load_id"], "Loaded", placed)
            else:
                add_row(ln, load["load_id"],
                        "PARTIAL -- %d of %d bundles did not fit (floor/stack limits); "
                        "re-plan or move to another load" % (miss, ln["bundles"]), placed)
    for ln in unassigned:
        add_row(ln, "-", "UNASSIGNED (no truck met minimum / fleet exhausted)", 0)

    sheet_style_print(ws, len(headers))
    ws.freeze_panes = "A2"
    return ws


DAY_ABBR = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def default_schedule_cfg():
    return {
        "offset_days": 2,      # load building today -> loading in N days
        "duration_h": 2.0,     # how long loading one truck takes
        "spacing_h": 2.0,      # start of one load to start of the next
        "deliver_start": "08:00", "deliver_end": "17:00",
        "shifts": {
            "WSM": {"weekday_day": "06:30-16:15", "weekday_night": "18:30-04:15",
                    "sat_day": "", "sat_night": "", "sun_day": "", "sun_night": "",
                    "skip_night_days": ["Thu"]},
            "LSM": {"weekday_day": "00:00-24:00", "weekday_night": "",
                    "sat_day": "00:00-24:00", "sat_night": "",
                    "sun_day": "00:00-24:00", "sun_night": "",
                    "skip_night_days": []},
        },
    }


def _parse_hhmm(s):
    from datetime import time
    s = str(s).strip()
    if not s:
        return None
    if s in ("24:00", "24h00", "2400"):
        return "MIDNIGHT_END"
    h, m = s.split(":")
    return time(int(h), int(m))


def _shift_windows(site_cfg, d):
    """Open loading windows (start_dt, end_dt) for calendar day d. A night
    window ending earlier than it starts crosses midnight into d+1."""
    from datetime import datetime, timedelta, time
    wd = d.weekday()
    if wd <= 4:
        keys = [("weekday_day", False), ("weekday_night", True)]
    elif wd == 5:
        keys = [("sat_day", False), ("sat_night", True)]
    else:
        keys = [("sun_day", False), ("sun_night", True)]
    wins = []
    for key, is_night in keys:
        raw = str(site_cfg.get(key) or "").strip()
        if not raw or "-" not in raw:
            continue
        if is_night and DAY_ABBR[wd] in site_cfg.get("skip_night_days", []):
            continue
        a, b = raw.split("-", 1)
        t1, t2 = _parse_hhmm(a), _parse_hhmm(b)
        if t1 is None or t2 is None or t1 == "MIDNIGHT_END":
            continue
        start = datetime.combine(d, t1)
        if t2 == "MIDNIGHT_END":
            end = datetime.combine(d + timedelta(days=1), time(0, 0))
        else:
            end = datetime.combine(d, t2)
            if end <= start:
                end += timedelta(days=1)   # crosses midnight
        wins.append((start, end))
    return sorted(wins)


def _next_loading_slot(site_cfg, not_before, dur_h):
    """Earliest start >= not_before where a full loading of dur_h hours fits
    inside one open shift window."""
    from datetime import timedelta
    dur = timedelta(hours=dur_h)
    for day_off in range(0, 60):
        d = (not_before.date() + timedelta(days=day_off))
        # include yesterday's overnight windows that spill into today
        wins = _shift_windows(site_cfg, d - timedelta(days=1)) + _shift_windows(site_cfg, d)
        for s, e in sorted(wins):
            start = max(s, not_before)
            if start + dur <= e:
                return start
    return not_before  # no configured windows at all -- schedule anyway


def write_transport_orders(wb, loads, sites, start_to_no=1000001, cfg=None):
    """TMS import tab. Per load (load_reference_group TRK001, TRK002...):
    one PICKUP transport order carrying every SKU line of the load
    (destination = the first/closest drop), then one TO per customer drop
    with that customer's lines. Each TO block = 1 main row, 4 window
    'placeholder' rows (delivery days +1..+4), then remaining SKU lines.
    Loading day = today + 2; the first load ships 07:00-09:00 and every
    following load shifts 2 hours later. TO numbers continue sequentially
    from start_to_no (set in the app to follow your TMS's numbering)."""
    from datetime import date, datetime, time, timedelta
    import load_builder as _lb

    cfg = cfg or default_schedule_cfg()
    ws = wb.create_sheet("Transport Orders")
    headers = ["Transport Order", "Trip", "load_reference_group", "Status", "Route Type",
               "Route Code", "Source Location", "Source Location Type", "Destination Location",
               "Destination Location Type", "Total Weight", "Total Volume", "Total Pallets",
               "Ship From", "Ship To", "Deliver From", "Deliver To",
               "Windows/Ship From", "Windows/Ship To", "Windows/Deliver From", "Windows/Deliver To",
               "Windows/Planning Priority", "Windows/Date From", "Windows/Date To", "External ID",
               "Lines/Demand Bucket Line/id", "Lines/Demand Bucket/External ID",
               "Lines/Demand Bucket Line/Line No", "Lines/Item Code", "Lines/Order ID",
               "Lines/Qty", "Lines/Total volume", "Lines/Total Weight"]
    ws.append(headers)
    for c in ws[1]:
        c.font = BOLD

    load_date = date.today() + timedelta(days=int(cfg["offset_days"]))
    to_no = int(start_to_no)
    del_start = _parse_hhmm(cfg["deliver_start"]) or time(8, 0)
    del_end = _parse_hhmm(cfg["deliver_end"]) or time(17, 0)
    site_cursor = {}   # per site: earliest datetime the next load may start

    def fmt(dtv):
        return dtv.strftime("%Y-%m-%d %H:%M")

    def line_cells(ln):
        db = getattr(_lb, "DEMAND_BUCKETS", {})
        b = db.get((_lb._order_digits(ln["sales_order"]), ln["sku"])) or db.get(ln["sku"]) or {}
        qty = int(round(ln["bundles"] * (ln.get("qty_nop") or 0))) or ln["bundles"]
        return [b.get("line_id", ""), b.get("bucket_ext", ""), b.get("line_no", ""),
                ln["sku"], "", qty, round(ln["m3"], 5), round(ln["bundle_kg"] * ln["bundles"], 3)]

    for li, load in enumerate(loads):
        site = load["site"]
        scfg = cfg["shifts"].get(site) or {"weekday_day": "00:00-24:00",
                                           "sat_day": "00:00-24:00", "sun_day": "00:00-24:00",
                                           "skip_night_days": []}
        not_before = site_cursor.get(site, datetime.combine(load_date, time(0, 0)))
        slot = _next_loading_slot(scfg, not_before, cfg["duration_h"])
        ship_from_dt = slot
        ship_to_dt = slot + timedelta(hours=cfg["duration_h"])
        site_cursor[site] = slot + timedelta(hours=cfg["spacing_h"])
        base = ship_from_dt.date()   # windows are anchored on the actual loading day
        sf, st_ = fmt(ship_from_dt), fmt(ship_to_dt)
        trk = "TRK%03d" % (li + 1)
        source = "SFP-%s - %s" % (site, sites.get(site, {}).get("name") or site)

        cust_lines, cust_dist = {}, {}
        for ln in load["lines"]:
            cust_lines.setdefault(ln["location_code"], []).append(ln)
            cust_dist[ln["location_code"]] = min(cust_dist.get(ln["location_code"], 1e18), ln["dist_km"])
        drops = sorted(cust_lines, key=lambda c: cust_dist[c])

        def emit(dest_code, dest_name, to_lines, ext_id):
            nonlocal to_no
            tot_kg = sum(l["bundle_kg"] * l["bundles"] for l in to_lines)
            tot_m3 = sum(l["m3"] for l in to_lines)
            tot_pal = sum(l["bundles"] for l in to_lines)
            lcells = [line_cells(l) for l in to_lines]
            # day-one delivery window: from loading end (or the delivery-window
            # opening, whichever is later) until the configured closing time
            day1_open = max(ship_to_dt, datetime.combine(base, del_start))
            day1_close = datetime.combine(base, del_end)
            if day1_open >= day1_close:
                day1_open = datetime.combine(base, del_start)
            ws.append(["TO%d" % to_no, "", trk, "Ready", "OUTBOUND", "",
                       source, "FACILITY", "%s - %s" % (dest_code, dest_name), "CUSTOMER",
                       round(tot_kg, 3), round(tot_m3, 5), tot_pal,
                       sf, st_, st_, fmt(datetime.combine(base + timedelta(days=4), del_end)),
                       sf, st_, fmt(day1_open), fmt(day1_close), "Must go",
                       base.isoformat(), base.isoformat(), ext_id] + lcells[0])
            for k in range(1, 5):
                d = base + timedelta(days=k)
                ws.append([""] * 17 + [sf, st_,
                                       fmt(datetime.combine(d, del_start)),
                                       fmt(datetime.combine(d, del_end)), "Must go",
                                       base.isoformat(), d.isoformat()] + [""] * 9)
            for lc in lcells[1:]:
                ws.append([""] * 25 + lc)
            to_no += 1

        first = drops[0]
        emit(first, cust_lines[first][0]["delivery_name"], load["lines"],
             cust_lines[first][0]["sales_order"])
        for code in drops:
            ls = cust_lines[code]
            emit(code, ls[0]["delivery_name"], ls, ls[0]["sales_order"])

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 19
    ws.freeze_panes = "A2"
    return ws


def _rounded_bundle(ax, x, y, w, h, accent, hatch):
    pad = min(w, h) * 0.05
    box = patches.FancyBboxPatch(
        (x + pad, y + pad), max(w - 2 * pad, 0.01), max(h - 2 * pad, 0.01),
        boxstyle="round,pad=0,rounding_size=%.4f" % (min(w, h) * 0.06),
        linewidth=1.3, edgecolor=accent, facecolor="white", hatch=hatch, zorder=3,
    )
    ax.add_patch(box)


def _badge(ax, x, y, w, h, pct, label):
    pct_clamped = max(0, min(pct, 100))
    color = "#2E7D32" if pct_clamped < 70 else ("#B8860B" if pct_clamped < 92 else "#B2323C")
    ax.add_patch(patches.FancyBboxPatch((x, y), w, h, boxstyle="round,pad=0,rounding_size=%.3f" % (h * 0.4),
                                         linewidth=1.1, edgecolor=color, facecolor="white", zorder=3))
    ax.text(x + w / 2, y + h / 2, "%s %.0f%%" % (label, pct_clamped), fontsize=6.2, va="center",
            ha="center", color=color, fontweight="bold", zorder=4)


TEXT_HALO = dict(boxstyle="round,pad=0.12", facecolor="white", edgecolor="none", alpha=0.88)


def draw_trailer(ax, trailer_info, title, style_map, seq_map):
    spec = trailer_info["spec"]
    length_cap, height_cap = spec["length_m"], spec["height_m"]
    top_pad = height_cap * 0.55
    ax.set_xlim(-0.6, length_cap + 0.3)
    ax.set_ylim(-0.55, height_cap * 2.2 + top_pad)
    ax.axis("off")

    ax.text(length_cap / 2, height_cap * 2.2 + top_pad * 0.68, title,
            fontsize=12, fontweight="bold", ha="center", color=NAVY)

    weight_pct = trailer_info["used_weight"] / (spec["weight_cap_t"] * 1000) * 100
    length_pct = trailer_info["used_length"] / spec["length_m"] * 100
    cube_pct = trailer_info["used_volume"] / trailer_info["cube_cap_m3"] * 100 if trailer_info["cube_cap_m3"] else 0
    badge_w = length_cap * 0.30
    badge_h = top_pad * 0.20
    badge_y = height_cap * 2.2 + top_pad * 0.10
    gap = (length_cap - 3 * badge_w) / 2
    _badge(ax, 0, badge_y, badge_w, badge_h, weight_pct, "WT")
    _badge(ax, badge_w + gap, badge_y, badge_w, badge_h, cube_pct, "CUBE")
    _badge(ax, 2 * (badge_w + gap), badge_y, badge_w, badge_h, length_pct, "FLOOR")

    for slot_idx, lane_label in ((0, "LEFT"), (1, "RIGHT")):
        y0 = height_cap * 1.2 if slot_idx == 1 else 0
        ax.add_patch(patches.FancyBboxPatch((0, y0), length_cap, height_cap,
                                             boxstyle="round,pad=0,rounding_size=0.08", linewidth=1.3,
                                             edgecolor=STEEL, facecolor="none", zorder=0))
        # height scale (m) from the truck's own dimensions, on the left edge
        hticks = list(range(0, int(height_cap) + 1))
        if height_cap - int(height_cap) > 1e-9:
            hticks.append(height_cap)
        for hval in hticks:
            y = y0 + hval
            ax.plot([-0.10, 0], [y, y], color="#888888", linewidth=0.8, zorder=1)
            ax.text(-0.16, y, "%g" % hval, fontsize=5.5, ha="right", va="center", color="#555555")
        # lane name rotated (reads bottom-to-top), centred on its lane box
        ax.text(-0.62, y0 + height_cap / 2, lane_label, rotation=90, va="center", ha="center",
                fontsize=8, color=STEEL, fontweight="bold")

    for p in trailer_info["placements"]:
        y_offset = height_cap * 1.2 if p["slot"] == 1 else 0
        y = y_offset + p["y"]
        accent, hatch = style_map.get(p.get("location_code"), (ACCENTS[0], HATCHES[0]))
        _rounded_bundle(ax, p["x"], y, p["bundle_length_m"], p["bundle_height_m"], accent, hatch)
        seq_no = seq_map.get(p.get("location_code"), (0, "", 0))[0]
        so_short = p["sales_order"].replace("SFP-", "").replace("-SO", "")
        h = p["bundle_height_m"]
        w = p["bundle_length_m"]
        if w < 1.05 or h < 0.16:
            label = "#%d" % seq_no
            fontsize = max(5.5, min(8.0, min(w, h) * 22))
        elif h >= 0.32:
            label = "#%d\nOrd %s\nSKU %s\n%s" % (seq_no, so_short, p["sku"], p["delivery_name"][:20])
            fontsize = max(4.0, min(6.0, h * 13))
        else:
            label = "#%d Ord %s | SKU %s\n%s" % (seq_no, so_short, p["sku"], p["delivery_name"][:20])
            fontsize = max(3.8, min(5.6, h * 15))
        clip_box = patches.Rectangle((p["x"], y), p["bundle_length_m"], p["bundle_height_m"], transform=ax.transData)
        ax.text(p["x"] + p["bundle_length_m"] / 2, y + p["bundle_height_m"] / 2, label,
                ha="center", va="center", fontsize=fontsize, color=INK, fontweight="bold",
                zorder=4, linespacing=1.0, clip_path=clip_box, clip_on=True, bbox=TEXT_HALO)

    # numeric length-axis scale along the bottom, plus a light chassis line
    ax.plot([-0.2, length_cap + 0.1], [-0.18, -0.18], color="#888888", linewidth=1.6, zorder=1)
    for wx in [length_cap * 0.18, length_cap * 0.5, length_cap * 0.82]:
        ax.add_patch(patches.Circle((wx, -0.18), 0.10, facecolor="none", edgecolor="#333333", linewidth=1.2, zorder=1))
    tick_step = 1 if length_cap <= 8 else 2
    tick = 0
    while tick <= length_cap + 1e-6:
        ax.plot([tick, tick], [-0.26, -0.20], color="#888888", linewidth=1.0, zorder=1)
        ax.text(tick, -0.48, "%g" % tick, fontsize=6.5, ha="center", color="#555555")
        tick += tick_step


def _draw_top_legend(fig, load, style_map, seq_map, rect):
    ax = fig.add_axes(rect)
    ax.axis("off")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    items = sorted(seq_map.items(), key=lambda kv: kv[1][0])
    n = max(len(items), 1)
    per_row = min(n, 5)
    col_w = 1.0 / per_row
    for i, (code, (seq_no, name, dist)) in enumerate(items):
        row = i // per_row
        col = i % per_row
        x = col * col_w
        y = 0.88 - row * 0.5
        accent, hatch = style_map.get(code, (ACCENTS[0], HATCHES[0]))
        ax.add_patch(patches.FancyBboxPatch((x, y - 0.12), 0.028, 0.20, boxstyle="round,pad=0,rounding_size=0.01",
                                             linewidth=1.0, edgecolor=accent, facecolor="white", hatch=hatch,
                                             clip_on=False))
        ax.text(x + 0.045, y - 0.02, "#%d  %s (%.0f km)" % (seq_no, (name or code)[:24], dist),
                fontsize=6.6, va="center", color="#333333")


def write_schematics_pdf(loads, path):
    n_total = len(loads)
    with PdfPages(path) as pdf:
        for page_no, load in enumerate(loads, start=1):
            fig = plt.figure(figsize=(11.7, 8.3))
            fig.patch.set_facecolor("white")

            header_ax = fig.add_axes([0, 0.93, 1, 0.07])
            header_ax.axis("off")
            header_ax.add_patch(patches.Rectangle((0, 0), 1, 0.05, transform=header_ax.transAxes,
                                                   facecolor=NAVY, edgecolor="none"))
            header_ax.text(0.01, 0.62, "LOAD %s" % load["load_id"], fontsize=17, fontweight="bold",
                            color=NAVY, va="center", transform=header_ax.transAxes)
            n_orders = len({ln["sales_order"] for ln in load["lines"]})
            n_skus = len({ln["sku"] for ln in load["lines"]})
            total_bundles = sum(ln["bundles"] for ln in load["lines"])
            placed_bundles = total_bundles - len(load["pack_leftover"])
            subtitle = ("Site: %s   |   %s   |   Truck: %s   |   %.1f m3, %.0f kg   |   "
                        "%s orders, %s SKUs, %s drops, %s lines   |   %d/%d bundles placed" % (
                load["site"], group_label(load["group"]), load["truck_type"], load["total_m3"], load["total_kg"],
                n_orders, n_skus, load["n_customers"], len(load["lines"]), placed_bundles, total_bundles))
            header_ax.text(0.01, 0.22, subtitle, fontsize=9, color="#444444", va="center",
                            transform=header_ax.transAxes)

            seq_map = _drop_sequence(load)
            style_map = _customer_style_map(load)
            _draw_top_legend(fig, load, style_map, seq_map, [0.01, 0.80, 0.98, 0.12])

            plot_area_width = 0.98
            names = list(load["packing"].keys())
            spans = [load["packing"][name]["spec"]["length_m"] + 0.9 for name in names]
            total_span = sum(spans)
            x_cursor = 0.01
            trailer_top = 0.70
            for i, name in enumerate(names):
                w = plot_area_width * spans[i] / total_span
                ax = fig.add_axes([x_cursor, 0.08, w - 0.02, trailer_top])
                draw_trailer(ax, load["packing"][name], "%s TRAILER" % name.upper(), style_map, seq_map)
                x_cursor += w

            footer_ax = fig.add_axes([0, 0, 1, 0.03])
            footer_ax.axis("off")
            footer_ax.text(0.99, 0.5, "Page %d of %d" % (page_no, n_total), fontsize=7.5, color="#999999",
                            ha="right", va="center", transform=footer_ax.transAxes)
            footer_ax.text(0.5, 0.5, "Build %s" % APP_VERSION, fontsize=7.5, color="#999999",
                           ha="center", va="center")
            footer_ax.text(0.01, 0.5, "Generated load-building schematic -- verify against physical stock before dispatch",
                            fontsize=7, color="#AAAAAA", ha="left", va="center", transform=footer_ax.transAxes)

            pdf.savefig(fig, orientation="landscape")
            plt.close(fig)


if __name__ == "__main__":
    loads, unassigned, fleet_left, lines, sites = build_all()

    wb = openpyxl.Workbook()
    write_loads_summary(wb, loads, unassigned)
    write_orders_summary(wb, loads, unassigned)
    write_transport_orders(wb, loads, sites)
    wb.save("Load Building Output.xlsx")
    print("Wrote Load Building Output.xlsx")

    write_schematics_pdf(loads, "Load Schematics.pdf")
    print("Wrote Load Schematics.pdf")

    print("Summary: %d loads, %d unassigned lines, fleet left: %s" % (len(loads), len(unassigned), fleet_left))
